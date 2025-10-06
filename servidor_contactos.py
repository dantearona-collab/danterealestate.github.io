from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os


app = Flask(__name__)
CORS(app)  # Permitir peticiones desde el navegador

# Nombre del archivo Excel
EXCEL_FILE = 'contactos_dante_propiedades.xlsx'

def init_excel():
    """Inicializa el archivo Excel si no existe"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contactos"
        # Crear encabezados
        ws['A1'] = 'Fecha/Hora'
        ws['B1'] = 'Nombre'
        ws['C1'] = 'Firma'
        ws['D1'] = 'Teléfono'
        ws['E1'] = 'Propiedad'
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        wb.save(EXCEL_FILE)
        print(f"✓ Archivo {EXCEL_FILE} creado exitosamente")
    else:
        print(f"✓ Archivo {EXCEL_FILE} encontrado")

def guardar_contacto(nombre, firma, telefono, propiedad="DESTACADA0"):
    """Guarda los datos del contacto en el archivo Excel"""
    try:
        # Cargar el archivo Excel existente
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Obtener la siguiente fila vacía
        next_row = ws.max_row + 1
        
        # Agregar los datos
        ws[f'A{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'B{next_row}'] = nombre
        ws[f'C{next_row}'] = firma if firma else '-'
        ws[f'D{next_row}'] = telefono
        ws[f'E{next_row}'] = propiedad
        
        # Guardar el archivo
        wb.save(EXCEL_FILE)
        print(f"✓ Contacto guardado: {nombre} - {telefono}")
        return True
    except Exception as e:
        print(f"✗ Error al guardar: {str(e)}")
        return False

@app.route('/guardar_contacto', methods=['POST', 'OPTIONS'])
def guardar_contacto_route():
    """Endpoint para guardar los datos del formulario"""
    
    # Manejar preflight CORS
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'message': 'No se recibieron datos'
            }), 400
        
        nombre = data.get('nombre', '').strip()
        firma = data.get('firma', '').strip()
        telefono = data.get('telefono', '').strip()
        propiedad = data.get('propiedad', 'DESTACADA0')
        
        # Validar que los campos requeridos no estén vacíos
        if not nombre or not telefono:
            return jsonify({
                'success': False,
                'message': 'El nombre y teléfono son obligatorios'
            }), 400
        
        # Guardar en Excel
        if guardar_contacto(nombre, firma, telefono, propiedad):
            return jsonify({
                'success': True,
                'message': 'Datos guardados correctamente en Excel'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Error al guardar los datos'
            }), 500
            
    except Exception as e:
        print(f"Error en el servidor: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error en el servidor: {str(e)}'
        }), 500

@app.route('/estadisticas')
def estadisticas():
    """Muestra estadísticas de contactos guardados"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            total = ws.max_row - 1  # Resta la fila de encabezados
            return jsonify({
                'success': True,
                'total_contactos': max(0, total),
                'archivo': EXCEL_FILE
            })
        return jsonify({
            'success': False, 
            'message': 'No hay datos todavía'
        })
    except Exception as e:
        return jsonify({
            'success': False, 
            'message': str(e)
        })

@app.route('/test')
def test():
    """Endpoint de prueba"""
    return jsonify({
        'success': True,
        'message': 'Servidor funcionando correctamente',
        'excel_existe': os.path.exists(EXCEL_FILE)
    })

if __name__ == '__main__':
    print("=" * 60)
    print("   SISTEMA DE CONTACTOS - DANTE PROPIEDADES")
    print("=" * 60)
    print()
    
    # Inicializar el archivo Excel
    init_excel()
    
    print()
    print("🚀 Servidor iniciando...")
    print(f"📋 Archivo Excel: {EXCEL_FILE}")
    print(f"📂 Ruta completa: {os.path.abspath(EXCEL_FILE)}")
    print("🌐 URL local: http://localhost:5000")
    print("🌐 Test: http://localhost:5000/test")
    print("📊 Estadísticas: http://localhost:5000/estadisticas")
    print()
    print("✅ El servidor está listo para recibir formularios")
    print("⚠️  Presiona Ctrl+C para detener el servidor")
    print("=" * 60)
    print()
    
    # Iniciar el servidor Flask
    app.run(debug=True, host='0.0.0.0', port=5000)