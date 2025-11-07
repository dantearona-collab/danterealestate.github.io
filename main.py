# pip install -r requirements.txt

import sys
import os
import json
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime
import mimetypes

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": ["null", "http://dantepropiedades.com.ar", "http://www.dantepropiedades.com.ar", "https://dantepropiedades.com.ar", "https://www.dantepropiedades.com.ar", "http://dantepropiedades.com", "https://danterealestate-github-io.onrender.com"]}})

# --- Excel Contact Logic ---
EXCEL_FILE = 'contactos_dante_propiedades.xlsx'

def safe_print(message):
    """Función segura para imprimir sin problemas de codificación"""
    safe_message = message.encode('ascii', 'ignore').decode('ascii')
    print(safe_message)

def init_excel():
    """Initializes the Excel file if it does not exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contactos"
        ws['A1'] = 'Fecha/Hora'
        ws['B1'] = 'Nombre'
        ws['C1'] = 'Firma'
        ws['D1'] = 'Teléfono'
        ws['E1'] = 'Propiedad'
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        wb.save(EXCEL_FILE)
        safe_print(f"SUCCESS: Archivo {EXCEL_FILE} creado exitosamente")
    else:
        safe_print(f"INFO: Archivo {EXCEL_FILE} encontrado")

def serve_static_file(filename):
    """Sirve archivos estáticos desde la raíz del repositorio - VERSIÓN OPTIMIZADA"""
    try:
        # Buscar el archivo en la raíz del repositorio
        file_path = os.path.join(os.getcwd(), filename)
        
        safe_print(f"🔍 Buscando archivo: {file_path}")
        
        if os.path.exists(file_path):
            try:
                # Obtener el tipo MIME correcto
                mime_type, _ = mimetypes.guess_type(filename)
                if not mime_type:
                    if filename.endswith('.css'):
                        mime_type = 'text/css'
                    elif filename.endswith('.js'):
                        mime_type = 'application/javascript'
                    elif filename.endswith('.json'):
                        mime_type = 'application/json'
                    elif filename.endswith('.html'):
                        mime_type = 'text/html'
                    else:
                        mime_type = 'text/plain'
                
                safe_print(f"✅ Archivo encontrado: {file_path}")
                safe_print(f"📏 Tamaño: {os.path.getsize(file_path)} bytes")
                safe_print(f"🎯 Tipo MIME: {mime_type}")
                
                # Leer el archivo
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                if filename.endswith('.css'):
                    safe_print(f"✅ Sirviendo {filename} como CSS (Cache: 1h)")
                    return content, 200, {'Content-Type': 'text/css', 'Cache-Control': 'public, max-age=3600'}
                elif filename.endswith('.js'):
                    safe_print(f"✅ Sirviendo {filename} como JavaScript (Cache: 1h)")
                    return content, 200, {'Content-Type': 'application/javascript', 'Cache-Control': 'public, max-age=3600'}
                elif filename.endswith('.json'):
                    safe_print(f"✅ Sirviendo {filename} como JSON (Cache: 1h)")
                    return content, 200, {'Content-Type': 'application/json', 'Cache-Control': 'public, max-age=3600'}
                elif filename.endswith('.html'):
                    safe_print(f"✅ Sirviendo {filename} como HTML (Cache: 1h)")
                    return content, 200, {'Content-Type': 'text/html', 'Cache-Control': 'public, max-age=3600'}
                else:
                    safe_print(f"✅ Sirviendo {filename} como {mime_type}")
                    return content, 200, {'Content-Type': mime_type}
                    
            except Exception as e:
                safe_print(f"❌ Error leyendo {filename}: {str(e)}")
                return jsonify({"error": f"Error leyendo archivo: {str(e)}"}), 500
        else:
            safe_print(f"❌ Archivo NO encontrado: {file_path}")
            return jsonify({"error": f"Archivo {filename} no encontrado en {os.getcwd()}"}), 404
            
    except Exception as e:
        safe_print(f"❌ Error sirviendo {filename}: {str(e)}")
        return jsonify({"error": f"Error sirviendo archivo: {str(e)}"}), 500

@app.route('/test-css')
def test_css():
    """Servir página de prueba CSS con diagnóstico"""
    test_html = '''<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Test CSS - Propiedades Dante</title>
    <link rel="stylesheet" href="app.css">
    <style>
        .test-container { max-width: 1200px; margin: 0 auto; padding: 20px; }
        .test-section { background: #f0f0f0; border: 2px solid #333; margin: 20px 0; padding: 20px; }
        .highlight { background: #ffff00; padding: 10px; border: 2px solid red; }
        .btn-test { background: blue; color: white; padding: 10px 20px; border: none; cursor: pointer; }
        .btn-test:hover { background: darkblue; }
    </style>
</head>
<body>
    <div class="test-container">
        <h1>🔥 DIAGNÓSTICO CSS - Propiedades Dante</h1>
        
        <div class="test-section">
            <h2>✅ Test 1: HTML Básico</h2>
            <p>Si ves este texto con formato básico, el HTML funciona correctamente</p>
        </div>
        
        <div class="highlight">
            <h2>🎯 Test 2: CSS Inline</h2>
            <p>Si este recuadro es <strong>AMARILLO con borde ROJO</strong>, el CSS inline funciona</p>
        </div>
        
        <div class="test-section">
            <h2>📋 Test 3: CSS Externo (app.css)</h2>
            <p>Si esta sección tiene <strong>background gris claro con borde</strong>, el archivo app.css se está aplicando</p>
            <p><strong>Tamaño esperado:</strong> 49,021 bytes</p>
        </div>
        
        <div class="test-section">
            <h2>🔧 Test 4: JavaScript</h2>
            <button class="btn-test" onclick="testJavaScript()">Probar JavaScript</button>
            <div id="js-result"></div>
        </div>
        
        <div class="test-section">
            <h2>🔍 Diagnóstico de Carga de Recursos</h2>
            <div id="resource-diagnosis">Cargando diagnóstico...</div>
        </div>
        
        <div class="test-section">
            <h2>📊 Resultados</h2>
            <ul>
                <li>✅ <strong>SI tienes formato:</strong> CSS funciona - Problema con HTML original</li>
                <li>❌ <strong>SI NO tienes formato:</strong> Problema con carga de recursos</li>
            </ul>
        </div>
    </div>
    
    <script>
        function testJavaScript() {
            document.getElementById('js-result').innerHTML = '✅ <strong style="color: green;">JavaScript funciona correctamente!</strong>';
        }
        
        // Diagnóstico automático de recursos
        function diagnoseResources() {
            const diagnosis = document.getElementById('resource-diagnosis');
            
            setTimeout(function() {
                // Verificar CSS
                const cssTest = getComputedStyle(document.querySelector('.test-section'));
                const hasCSS = cssTest.backgroundColor === 'rgb(240, 240, 240)' || cssTest.backgroundColor === '#f0f0f0';
                
                // Verificar JavaScript
                const hasJS = typeof testJavaScript === 'function';
                
                // Verificar fuentes
                const font = getComputedStyle(document.body).fontFamily;
                const hasCustomFont = font.includes('Fira Sans') || font.includes('sans-serif');
                
                let result = '<h3>Diagnóstico de Recursos:</h3><ul>';
                result += '<li>' + (hasCSS ? '✅' : '❌') + ' <strong>CSS:</strong> ' + (hasCSS ? 'Aplicándose correctamente' : 'NO se está aplicando') + '</li>';
                result += '<li>' + (hasJS ? '✅' : '❌') + ' <strong>JavaScript:</strong> ' + (hasJS ? 'Funcionando' : 'No funciona') + '</li>';
                result += '<li>' + (hasCustomFont ? '✅' : '❌') + ' <strong>Fuentes:</strong> ' + (hasCustomFont ? 'Cargando (Fira Sans)' : 'Fallback (sans-serif)') + '</li>';
                result += '</ul>';
                
                diagnosis.innerHTML = result;
            }, 1000);
        }
        
        diagnoseResources();
    </script>
</body>
</html>'''
    return test_html, 200, {'Content-Type': 'text/html'}

@app.route('/debug')
def debug_info():
    """Información de debug sobre archivos disponibles"""
    debug_info = {
        "current_directory": os.getcwd(),
        "files_in_directory": sorted(os.listdir('.')),
        "app_css_exists": os.path.exists('app.css'),
        "app_js_exists": os.path.exists('app.js'),
        "index_html_exists": os.path.exists('index.html'),
        "test_css_exists": True,  # Esta página siempre existe
        "propiedades_json_exists": os.path.exists('propiedades.json'),
        "main_py_exists": os.path.exists('main.py')
    }
    return jsonify(debug_info)

@app.route('/<path:filename>')
def serve_static(filename):
    """Ruta genérica para servir archivos estáticos"""
    return serve_static_file(filename)

@app.route('/api/guardar_contacto', methods=['POST', 'OPTIONS'])
@app.route('/guardar_contacto', methods=['POST', 'OPTIONS'])
def guardar_contacto_route():
    """Endpoint to save form data"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        data = request.get_json()
        init_excel()
        
        fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        nombre = data.get('nombre', '').strip()
        firma = data.get('firma', '').strip()
        telefono = data.get('telefono', '').strip()
        propiedad = data.get('propiedad', '').strip()
        
        if not nombre and not telefono:
            return jsonify({'error': 'Se requiere al menos nombre o teléfono'}), 400
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        next_row = ws.max_row + 1
        
        ws[f'A{next_row}'] = fecha_hora
        ws[f'B{next_row}'] = nombre
        ws[f'C{next_row}'] = firma
        ws[f'D{next_row}'] = telefono
        ws[f'E{next_row}'] = propiedad
        
        wb.save(EXCEL_FILE)
        safe_print(f"SUCCESS: Contacto guardado - {nombre} - {telefono} - {fecha_hora}")
        
        return jsonify({'message': 'Contacto guardado exitosamente'}), 200
        
    except Exception as e:
        safe_print(f"ERROR: {str(e)}")
        return jsonify({'error': f'Error interno: {str(e)}'}), 500

@app.route("/")
def index():
    """Servir la página principal ORIGINAL (no la de prueba)"""
    try:
        return serve_static_file('index.html')
    except Exception as e:
        safe_print(f"Error sirviendo index.html: {str(e)}")
        return jsonify({"error": f"Error sirviendo index.html: {str(e)}"}), 500

# --- API Endpoints ---

@app.route("/api/properties/search", methods=["GET"])
def search_properties():
    """Búsqueda avanzada de propiedades con filtros flexibles"""
    safe_print("--- Nueva Búsqueda ---")
    try:
        with open('propiedades.json', 'r', encoding='utf-8') as f:
            properties = json.load(f)
    except FileNotFoundError:
        return jsonify({"error": "El archivo propiedades.json no fue encontrado."}), 404
    except json.JSONDecodeError:
        return jsonify({"error": "El archivo propiedades.json no tiene un formato JSON válido."}), 500

    # Parámetros de búsqueda
    ope = request.args.get("ope")
    tipo = request.args.get("tipo")
    loc = request.args.get("loc")
    cod = request.args.get("cod")
    precio_min = request.args.get("precio_min")
    precio_max = request.args.get("precio_max")
    ambientes = request.args.get("ambientes")

    safe_print(f"Parámetros recibidos: ope={ope}, tipo={tipo}, loc={loc}, cod={cod}, precio_min={precio_min}, precio_max={precio_max}, ambientes={ambientes}")

    # Mapeo de operaciones
    operation_map = {
        "V": "venta",
        "A": "alquiler", 
        "T": "alquiler temporal"
    }
    
    if ope in operation_map:
        ope = operation_map[ope]

    filtered_properties = []
    for prop in properties:
        # Match operation
        if ope and (prop.get('operacion') is None or prop.get('operacion').lower() != ope.lower()):
            continue
        
        # Match property type
        if tipo and (prop.get('tipo') is None or prop.get('tipo').lower() != tipo.lower()):
            continue

        # Match neighborhood (busqueda flexible por substring)
        if loc and (prop.get('barrio') is None or loc.lower() not in prop.get('barrio', '').lower()):
            continue

        # Match code
        if cod and (prop.get('id_temporal') is None or cod.lower() not in prop.get('id_temporal', '').lower()):
            continue

        # Match price range
        if precio_min and prop.get('precio'):
            if float(prop.get('precio', 0)) < float(precio_min):
                continue
                
        if precio_max and prop.get('precio'):
            if float(prop.get('precio', 0)) > float(precio_max):
                continue

        # Match ambientes
        if ambientes and prop.get('ambientes'):
            if int(prop.get('ambientes', 0)) < int(ambientes):
                continue

        filtered_properties.append(prop)
    
    safe_print(f"Propiedades encontradas: {len(filtered_properties)}")
    
    # Agregar información de metadatos
    response = {
        "properties": filtered_properties,
        "total_found": len(filtered_properties),
        "filters_applied": {
            "operacion": ope,
            "tipo": tipo,
            "barrio": loc,
            "codigo": cod,
            "precio_min": precio_min,
            "precio_max": precio_max,
            "ambientes_min": ambientes
        }
    }
    
    return jsonify(response)

@app.route("/api/properties/filter-options", methods=["GET"])
def get_filter_options():
    """Obtiene las opciones disponibles para los filtros"""
    try:
        with open('propiedades.json', 'r', encoding='utf-8') as f:
            properties = json.load(f)
    except FileNotFoundError:
        return jsonify({"error": "El archivo propiedades.json no fue encontrado."}), 404

    # Obtener barrios únicos
    barrios = []
    tipos = []
    operaciones = []
    
    for prop in properties:
        # Barrios
        if prop.get('barrio') and prop['barrio'] not in barrios:
            barrios.append(prop['barrio'])
        
        # Tipos
        if prop.get('tipo') and prop['tipo'] not in tipos:
            tipos.append(prop['tipo'])
        
        # Operaciones
        if prop.get('operacion') and prop['operacion'] not in operaciones:
            operaciones.append(prop['operacion'])
    
    return jsonify({
        "barrios": sorted(barrios),
        "tipos": sorted(tipos),
        "operaciones": sorted(operaciones)
    })

@app.route("/api/properties/stats", methods=["GET"])
def get_property_stats():
    """Obtiene estadísticas generales de las propiedades"""
    try:
        with open('propiedades.json', 'r', encoding='utf-8') as f:
            properties = json.load(f)
    except FileNotFoundError:
        return jsonify({"error": "El archivo propiedades.json no fue encontrado."}), 404

    stats = {
        "total_properties": len(properties),
        "by_operacion": {},
        "by_tipo": {},
        "by_barrio": {},
        "price_ranges": {
            "min": None,
            "max": None,
            "avg": None
        }
    }
    
    precios = []
    
    for prop in properties:
        # Contar por operación
        op = prop.get('operacion', 'No especificado')
        stats["by_operacion"][op] = stats["by_operacion"].get(op, 0) + 1
        
        # Contar por tipo
        tipo = prop.get('tipo', 'No especificado')
        stats["by_tipo"][tipo] = stats["by_tipo"].get(tipo, 0) + 1
        
        # Contar por barrio
        barrio = prop.get('barrio', 'No especificado')
        stats["by_barrio"][barrio] = stats["by_barrio"].get(barrio, 0) + 1
        
        # Recopilar precios
        if prop.get('precio'):
            precios.append(float(prop['precio']))
    
    # Calcular rangos de precios
    if precios:
        stats["price_ranges"]["min"] = min(precios)
        stats["price_ranges"]["max"] = max(precios)
        stats["price_ranges"]["avg"] = sum(precios) / len(precios)
    
    return jsonify(stats)

if __name__ == '__main__':
    init_excel()  # Initialize Excel file
    safe_print("INFO: Iniciando servidor Flask...")
    safe_print("INFO: Excel inicializado")
    safe_print("INFO: Servidor corriendo en: http://127.0.0.1:5000")
    app.run(host='0.0.0.0', port=5000)