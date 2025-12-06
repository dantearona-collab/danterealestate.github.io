# -*- coding: utf-8 -*-
"""
🚀 SISTEMA DE ALMACENAMIENTO EXCEL - BACKEND COMPLETO
================================================================

Este servidor Python recibe los datos del formulario y los almacena automáticamente
en Excel y CSV. Usa openpyxl para mejor compatibilidad con Python 3.13.
"""

from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import json
import os
from datetime import datetime, timedelta
import csv
import logging
import shutil
from pathlib import Path

# Asegurar que existe el directorio de logs
os.makedirs('data', exist_ok=True)

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('data/sistema-formularios.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

class ExcelStorageManager:
    """
    📊 Gestor completo de almacenamiento en Excel usando openpyxl
    """
    
    def __init__(self, base_path='data'):
        self.base_path = Path(base_path)
        self.excel_path = self.base_path / 'excel' / 'consultas.xlsx'
        self.csv_path = self.base_path / 'excel' / 'consultas.csv'
        self.backup_path = self.base_path / 'backups'
        
        # Crear directorios si no existen
        self._crear_estructura_directorios()
        
        # Configuración de columnas
        self.columnas = [
            'Fecha', 'Hora', 'Timestamp', 'Nombre', 'Email', 'Teléfono', 
            'Interés', 'Presupuesto', 'Mensaje', 'Página', 'IP', 'User_Agent',
            'Estado', 'Notas'
        ]
        
        # Inicializar archivos
        self._inicializar_archivos()
        
        logging.info(f"✅ Sistema de almacenamiento inicializado en: {self.base_path}")
    
    def _crear_estructura_directorios(self):
        """Crear estructura de directorios necesaria"""
        directorios = [
            self.base_path / 'excel',
            self.base_path / 'backups',
            self.base_path / 'temp'
        ]
        
        for directorio in directorios:
            directorio.mkdir(parents=True, exist_ok=True)
    
    def _inicializar_archivos(self):
        """Inicializar archivos Excel y CSV con encabezados"""
        # Crear Excel si no existe
        if not self.excel_path.exists():
            wb = Workbook()
            
            # Hoja principal - Consultas
            ws_main = wb.active
            ws_main.title = 'Consultas'
            for col_idx, col_name in enumerate(self.columnas, 1):
                ws_main.cell(row=1, column=col_idx, value=col_name)
            
            # Hoja de backup
            ws_backup = wb.create_sheet('Backup')
            for col_idx, col_name in enumerate(self.columnas, 1):
                ws_backup.cell(row=1, column=col_idx, value=col_name)
            
            # Hoja de estadísticas
            ws_stats = wb.create_sheet('Estadísticas')
            ws_stats.cell(row=1, column=1, value='Métrica')
            ws_stats.cell(row=1, column=2, value='Valor')
            
            wb.save(str(self.excel_path))
            logging.info(f"✅ Archivo Excel creado: {self.excel_path}")
        
        # Crear CSV con encabezados
        if not self.csv_path.exists():
            with open(self.csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(self.columnas)
            logging.info(f"✅ Archivo CSV creado: {self.csv_path}")
    
    def añadir_consulta(self, datos_formulario):
        """
        📝 Añadir nueva consulta al sistema de almacenamiento
        """
        try:
            # Preparar datos completos
            timestamp = datetime.now()
            
            consulta_completa = {
                'Fecha': timestamp.strftime('%d/%m/%Y'),
                'Hora': timestamp.strftime('%H:%M:%S'),
                'Timestamp': timestamp.isoformat(),
                'Nombre': datos_formulario.get('nombre', ''),
                'Email': datos_formulario.get('email', ''),
                'Teléfono': datos_formulario.get('telefono', ''),
                'Interés': datos_formulario.get('interes', ''),
                'Presupuesto': datos_formulario.get('presupuesto', ''),
                'Mensaje': datos_formulario.get('mensaje', ''),
                'Página': datos_formulario.get('pagina', 'Desconocida'),
                'IP': request.remote_addr if request else 'N/A',
                'User_Agent': request.headers.get('User-Agent', 'N/A') if request else 'N/A',
                'Estado': 'Nueva',
                'Notas': ''
            }
            
            # Guardar en Excel (sheet principal)
            self._guardar_en_excel(consulta_completa)
            
            # Guardar en CSV
            self._guardar_en_csv(consulta_completa)
            
            # Crear backup automático
            self._crear_backup_automatico()
            
            logging.info(f"✅ Consulta guardada: {consulta_completa['Nombre']} - {consulta_completa['Email']}")
            
            return {
                'success': True,
                'message': 'Consulta guardada correctamente',
                'timestamp': consulta_completa['Timestamp'],
                'file': str(self.excel_path)
            }
            
        except Exception as e:
            logging.error(f"❌ Error guardando consulta: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'message': 'Error al guardar consulta'
            }
    
    def _guardar_en_excel(self, consulta):
        """💾 Guardar en archivo Excel usando openpyxl"""
        try:
            # Cargar workbook existente
            wb = load_workbook(str(self.excel_path))
            
            # Obtener hoja Consultas
            if 'Consultas' in wb.sheetnames:
                ws = wb['Consultas']
            else:
                ws = wb.create_sheet('Consultas')
                # Escribir encabezados si es nueva
                for col_idx, col_name in enumerate(self.columnas, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
            
            # Encontrar última fila
            last_row = ws.max_row + 1
            
            # Escribir datos
            for col_idx, col_name in enumerate(self.columnas, 1):
                ws.cell(row=last_row, column=col_idx, value=consulta.get(col_name, ''))
            
            # Ajustar ancho de columnas
            self._ajustar_ancho_columnas(ws)
            
            # Actualizar hoja de backup
            self._actualizar_backup(wb, consulta)
            
            # Actualizar estadísticas
            self._actualizar_estadisticas(wb)
            
            # Guardar
            wb.save(str(self.excel_path))
            
            logging.info("✅ Guardado en Excel completado")
            
        except Exception as e:
            logging.error(f"❌ Error guardando en Excel: {str(e)}")
            raise
    
    def _ajustar_ancho_columnas(self, worksheet):
        """Ajustar ancho de columnas automáticamente"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _actualizar_backup(self, workbook, consulta):
        """Actualizar hoja de backup"""
        if 'Backup' not in workbook.sheetnames:
            ws_backup = workbook.create_sheet('Backup')
            for col_idx, col_name in enumerate(self.columnas, 1):
                ws_backup.cell(row=1, column=col_idx, value=col_name)
        else:
            ws_backup = workbook['Backup']
        
        last_row = ws_backup.max_row + 1
        for col_idx, col_name in enumerate(self.columnas, 1):
            ws_backup.cell(row=last_row, column=col_idx, value=consulta.get(col_name, ''))
    
    def _actualizar_estadisticas(self, workbook):
        """Actualizar hoja de estadísticas"""
        if 'Consultas' not in workbook.sheetnames:
            return
        
        ws = workbook['Consultas']
        total_rows = ws.max_row - 1  # Restar encabezado
        
        # Crear o limpiar hoja de estadísticas
        if 'Estadísticas' in workbook.sheetnames:
            ws_stats = workbook['Estadísticas']
            # Limpiar datos existentes (mantener encabezados)
            for row in ws_stats.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
        else:
            ws_stats = workbook.create_sheet('Estadísticas')
            ws_stats.cell(row=1, column=1, value='Métrica')
            ws_stats.cell(row=1, column=2, value='Valor')
        
        # Calcular estadísticas básicas
        fecha_hoy = datetime.now().strftime('%d/%m/%Y')
        consultas_hoy = 0
        intereses = {}
        presupuestos = {}
        
        # Analizar datos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # Fecha
            fecha_cell = row[0]
            if fecha_cell.value == fecha_hoy:
                consultas_hoy += 1
            
            # Interés (columna 7)
            interes_cell = row[6]
            if interes_cell.value:
                interes = str(interes_cell.value)
                intereses[interes] = intereses.get(interes, 0) + 1
            
            # Presupuesto (columna 8)
            presupuesto_cell = row[7]
            if presupuesto_cell.value:
                presupuesto = str(presupuesto_cell.value)
                presupuestos[presupuesto] = presupuestos.get(presupuesto, 0) + 1
        
        # Interés más común
        interes_comun = "N/A"
        if intereses:
            interes_comun = max(intereses, key=intereses.get)
        
        # Presupuesto más común
        presupuesto_comun = "N/A"
        if presupuestos:
            presupuesto_comun = max(presupuestos, key=presupuestos.get)
        
        # Última consulta
        ultima_fecha = "N/A"
        if total_rows > 0:
            ultima_fecha = ws.cell(row=ws.max_row, column=1).value or "N/A"
        
        # Escribir estadísticas
        estadisticas = [
            ('Total Consultas', total_rows),
            ('Consultas Hoy', consultas_hoy),
            ('Interés Más Común', interes_comun),
            ('Presupuesto Más Común', presupuesto_comun),
            ('Última Consulta', ultima_fecha),
            ('Última Actualización', datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        ]
        
        for idx, (metrica, valor) in enumerate(estadisticas, start=2):
            ws_stats.cell(row=idx, column=1, value=metrica)
            ws_stats.cell(row=idx, column=2, value=valor)
    
    def _guardar_en_csv(self, consulta):
        """📄 Guardar en archivo CSV"""
        try:
            with open(self.csv_path, 'a', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=self.columnas)
                writer.writerow(consulta)
            
            logging.info("✅ Guardado en CSV completado")
            
        except Exception as e:
            logging.error(f"❌ Error guardando en CSV: {str(e)}")
            raise
    
    def _crear_backup_automatico(self):
        """🗂️ Crear backup automático cada 50 consultas"""
        try:
            # Contar consultas actuales
            wb = load_workbook(str(self.excel_path), data_only=True)
            ws = wb['Consultas']
            total_consultas = ws.max_row - 1  # Restar encabezado
            
            if total_consultas > 0 and total_consultas % 50 == 0:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_file = self.backup_path / f'backup_consultas_{timestamp}.xlsx'
                
                # Copiar archivo completo
                shutil.copy2(self.excel_path, backup_file)
                
                logging.info(f"🗂️ Backup creado: {backup_file}")
                
        except Exception as e:
            logging.error(f"❌ Error creando backup: {str(e)}")
    
    def obtener_consultas(self, limite=100):
        """📋 Obtener últimas consultas"""
        try:
            wb = load_workbook(str(self.excel_path), data_only=True)
            
            if 'Consultas' not in wb.sheetnames:
                return []
            
            ws = wb['Consultas']
            
            # Obtener encabezados
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # Obtener datos (últimas 'limite' filas)
            consultas = []
            start_row = max(2, ws.max_row - limite + 1)
            
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
                consulta = {}
                for header, cell in zip(headers, row):
                    consulta[header] = cell.value
                consultas.append(consulta)
            
            return consultas
            
        except Exception as e:
            logging.error(f"❌ Error obteniendo consultas: {str(e)}")
            return []
    
    def exportar_resumen(self):
        """📊 Exportar resumen de estadísticas"""
        try:
            wb = load_workbook(str(self.excel_path), data_only=True)
            
            if 'Consultas' not in wb.sheetnames:
                return "No hay datos para exportar"
            
            ws = wb['Consultas']
            total_consultas = ws.max_row - 1
            
            if total_consultas == 0:
                return "No hay consultas registradas"
            
            # Contar consultas de hoy
            fecha_hoy = datetime.now().strftime('%d/%m/%Y')
            consultas_hoy = 0
            intereses = {}
            presupuestos = {}
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                fecha_cell = row[0]
                if fecha_cell.value == fecha_hoy:
                    consultas_hoy += 1
                
                interes_cell = row[6]
                if interes_cell.value:
                    interes = str(interes_cell.value)
                    intereses[interes] = intereses.get(interes, 0) + 1
                
                presupuesto_cell = row[7]
                if presupuesto_cell.value:
                    presupuesto = str(presupuesto_cell.value)
                    presupuestos[presupuesto] = presupuestos.get(presupuesto, 0) + 1
            
            # Top 5 intereses
            top_intereses = sorted(intereses.items(), key=lambda x: x[1], reverse=True)[:5]
            intereses_str = "\n".join([f"  • {k}: {v}" for k, v in top_intereses]) if top_intereses else "  No hay datos"
            
            # Top 5 presupuestos
            top_presupuestos = sorted(presupuestos.items(), key=lambda x: x[1], reverse=True)[:5]
            presupuestos_str = "\n".join([f"  • {k}: {v}" for k, v in top_presupuestos]) if top_presupuestos else "  No hay datos"
            
            resumen = f"""
📊 RESUMEN DE CONSULTAS - {datetime.now().strftime('%d/%m/%Y %H:%M')}

📈 ESTADÍSTICAS GENERALES:
• Total de consultas: {total_consultas}
• Consultas hoy: {consultas_hoy}
• Última consulta: {fecha_hoy if consultas_hoy > 0 else 'N/A'}

🎯 INTERESES MÁS CONSULTADOS (Top 5):
{intereses_str}

💰 PRESUPUESTOS MÁS CONSULTADOS (Top 5):
{presupuestos_str}

📁 ARCHIVOS:
• Excel: {self.excel_path}
• CSV: {self.csv_path}
• Backups: {self.backup_path}
            """
            
            return resumen
            
        except Exception as e:
            return f"Error generando resumen: {str(e)}"

# Crear aplicación Flask
app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

# Inicializar gestor de almacenamiento
storage_manager = ExcelStorageManager()

@app.route('/')
def home():
    """🏠 Página principal - Sirve index.html o muestra API"""
    try:
        # Verificar si existe index.html
        if os.path.exists('index.html'):
            return send_file('index.html')
        else:
            # Si no existe, mostrar info de API
            return jsonify({
                'message': '🚀 Sistema de Formularios con Almacenamiento Excel',
                'version': '1.3.0',
                'status': 'active',
                'note': 'index.html no encontrado en el servidor',
                'endpoints': {
                    '/health': 'GET - Estado del sistema',
                    '/api/guardar-contacto': 'POST - Guardar consulta de contacto',
                    '/api/obtener-consultas': 'GET - Obtener últimas consultas',
                    '/api/resumen': 'GET - Obtener resumen estadístico',
                    '/api/exportar-excel': 'GET - Descargar archivo Excel'
                }
            })
    except Exception as e:
        logging.error(f"❌ Error en ruta principal: {str(e)}")
        return jsonify({'error': str(e)}), 500



# 📁 SERVIR ARCHIVOS ESTÁTICOS
@app.route('/<path:filename>')
def serve_static(filename):
    """Servir archivos estáticos (CSS, JS, imágenes, etc.)"""
    try:
        # Lista de extensiones permitidas
        allowed_ext = {'.html', '.css', '.js', '.png', '.jpg', '.jpeg', 
                      '.gif', '.ico', '.svg', '.json', '.txt', '.woff', 
                      '.woff2', '.ttf', '.eot'}
        
        filepath = Path(filename)
        
        # Verificar si existe
        if not filepath.exists():
            return jsonify({'error': 'Archivo no encontrado'}), 404
        
        # Verificar extensión
        if filepath.suffix.lower() not in allowed_ext:
            return jsonify({'error': 'Tipo de archivo no permitido'}), 403
        
        return send_file(str(filepath))
        
    except Exception as e:
        logging.error(f"❌ Error sirviendo {filename}: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500

# 🎨 RUTAS ESPECÍFICAS PARA CARPETAS COMUNES
@app.route('/css/<path:filename>')
def serve_css(filename):
    """Servir archivos CSS"""
    return send_from_directory('css', filename)

@app.route('/js/<path:filename>')
def serve_js(filename):
    """Servir archivos JavaScript"""
    return send_from_directory('js', filename)

@app.route('/imgs/<path:filename>')
def serve_imgs(filename):
    """Servir imágenes"""
    return send_from_directory('imgs', filename)

@app.route('/img/<path:filename>')
def serve_img(filename):
    """Servir imágenes (alternativa)"""
    return send_from_directory('img', filename)










@app.route('/health')
def health_check():
    """🏥 Verificar estado del sistema"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'storage_path': str(storage_manager.excel_path),
        'files_exist': {
            'excel': storage_manager.excel_path.exists(),
            'csv': storage_manager.csv_path.exists()
        },
        'python_version': os.sys.version
    })

@app.route('/api/guardar-contacto', methods=['POST'])
def guardar_contacto():
    """💾 Guardar nueva consulta de contacto"""
    try:
        # Verificar que se envió JSON
        if not request.is_json:
            return jsonify({
                'success': False,
                'error': 'Se requiere contenido JSON'
            }), 400
        
        datos = request.get_json()
        
        # Validaciones básicas
        campos_requeridos = ['nombre', 'email', 'mensaje']
        for campo in campos_requeridos:
            if not datos.get(campo, '').strip():
                return jsonify({
                    'success': False,
                    'error': f'Campo requerido faltante: {campo}'
                }), 400
        
        # Validar email
        email = datos.get('email', '')
        if '@' not in email:
            return jsonify({
                'success': False,
                'error': 'Email inválido'
            }), 400
        
        # Añadir información de contexto
        datos['pagina'] = request.headers.get('Referer', 'Directo')
        
        # Guardar en almacenamiento
        resultado = storage_manager.añadir_consulta(datos)
        
        if resultado['success']:
            return jsonify(resultado), 200
        else:
            return jsonify(resultado), 500
            
    except Exception as e:
        logging.error(f"❌ Error en guardar-contacto: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Error interno del servidor'
        }), 500

@app.route('/api/obtener-consultas', methods=['GET'])
def obtener_consultas():
    """📋 Obtener últimas consultas"""
    try:
        limite = request.args.get('limite', 100, type=int)
        consultas = storage_manager.obtener_consultas(limite)
        
        return jsonify({
            'success': True,
            'consultas': consultas,
            'total': len(consultas)
        })
        
    except Exception as e:
        logging.error(f"❌ Error en obtener-consultas: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/resumen', methods=['GET'])
def obtener_resumen():
    """📊 Obtener resumen estadístico"""
    try:
        resumen = storage_manager.exportar_resumen()
        return jsonify({
            'success': True,
            'resumen': resumen,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        logging.error(f"❌ Error en resumen: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/exportar-excel', methods=['GET'])
def exportar_excel():
    """📊 Exportar archivo Excel"""
    try:
        if not storage_manager.excel_path.exists():
            return jsonify({
                'success': False,
                'error': 'No hay datos para exportar'
            }), 404
        
        # Leer archivo y devolver como descarga
        return send_file(
            str(storage_manager.excel_path),
            as_attachment=True,
            download_name=f'consultas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logging.error(f"❌ Error exportando Excel: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

if __name__ == '__main__':
    print("🚀 Iniciando Sistema de Formularios con Almacenamiento Excel (openpyxl)")
    print(f"📁 Archivos de datos en: {storage_manager.base_path}")
    print(f"📊 Excel: {storage_manager.excel_path}")
    print(f"📄 CSV: {storage_manager.csv_path}")
    print("🌐 Servidor disponible en: http://localhost:5000")
    print("=" * 60)
    
    # Usar puerto del entorno o 5000 por defecto
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=True, host='0.0.0.0', port=port)